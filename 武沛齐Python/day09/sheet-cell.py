import os
from openpyxl import load_workbook

# 1.获取work_book对象
file_path = os.path.join("files", "output.xlsx")
work_book_object = load_workbook(file_path)

# 2.获取所有的sheet对象（内部包含很多数据）
sheet_object_list = work_book_object.worksheets

# 3.获取某一个sheet对象：方式1:
sheet_object = sheet_object_list[0]


# 读取单元格对象的方式1：
# 4.读取sheet中的单元格，cell对象
cell_object1 = sheet_object.cell(1, 1)

# 5.读取cell对象中的文本或数值
print(cell_object1.value)


# 读取单元格对象的方式2：
cell_object2 = sheet_object["A1"]
print(cell_object2.value)


# 输出
-0.015709151
-0.015709151
