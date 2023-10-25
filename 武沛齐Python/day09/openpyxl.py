import os
from openpyxl import load_workbook

file_path = os.path.join("files", "output.xlsx")

# 1.获取work_book对象
work_book_object = load_workbook(file_path)

# 2.读取所有sheet的名字
data_list = work_book_object.sheetnames
print(data_list)

# 输出
['Sheet1', 'Sheet2']
