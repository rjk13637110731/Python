import os
from openpyxl import load_workbook

# 1.获取work_book对象
file_path = os.path.join("files", "output.xlsx")
work_book_object = load_workbook(file_path)

# 2.获取所有的sheet对象（内部包含很多数据）
sheet_object_list = work_book_object.worksheets

# 3.获取某一个sheet对象：方式1:
sheet_object = sheet_object_list[1]

# 4.读取指定列数据
column1 = []
column2 = []
column3 = []
for row in sheet_object.iter_rows(min_row=2):
    column1.append(row[0].value)
    column2.append(row[1].value)
    column3.append(row[2].value)
print(column1)
print(column2)
print(column3)


# 输出
['学生姓名', '我害怕', None, '大白菜', 'ff', None]
['学生ID', '14789456546546', '14789456546545', '14346546546541', None, None]
[None, None, None, None, None, None]
