import os
from openpyxl import load_workbook

# 1.获取work_book对象
file_path = os.path.join("files", "output.xlsx")
work_book_object = load_workbook(file_path)

# 2.获取所有的sheet对象（内部包含很多数据）
sheet_object_list = work_book_object.worksheets

# 3.获取某一个sheet对象：方式1:
sheet_object = sheet_object_list[1]

# 4.获取合并单元格
cell_1_object = sheet_object.cell(1, 1)  # Cell对象
print(cell_1_object)
print(cell_1_object.value)

cell_2_object = sheet_object.cell(1, 2)  # MergedCell对象，默认是没有值的
print(cell_2_object)
print(cell_2_object.value)

cell_3_object = sheet_object.cell(3, 1)  # Cell对象，有值有内容
print(cell_3_object)
print(cell_3_object.value)

cell_4_object = sheet_object.cell(4, 1)  # MergedCell对象，默认是没有值的
print(cell_4_object)
print(cell_4_object.value)

# 输出
<Cell 'Sheet2'.A1>
用户信息
<MergedCell 'Sheet2'.B1>
None
<Cell 'Sheet2'.A3>
我害怕
<MergedCell 'Sheet2'.A4>
None
