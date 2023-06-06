import openpyxl


name2Age = {
    '张飞' :  38,
    '赵云' :  27,
    '许褚' :  36,
    '典韦' :  38,
    '关羽' :  39,
    '黄忠' :  49,
    '徐晃' :  43,
    '马超' :  23,
}

book = openpyxl.Workbook()
sh = book.active

sh.title = '年龄表'

sh['A1'] = '姓名'
sh['B1'] = '年龄'

row = 2

for name,age in name2Age.items():
    sh.cell(row,1).value = name
    sh.cell(row,2).value = age
    row += 1
book.save('个人信息表.xlsx')
